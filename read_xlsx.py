import openpyxl,os

def ab_order(num):
    count=1;comp=25
    while num>comp:
        count+=1;comp+=26**count
    re='';n=num-comp+26**count;c=0
    while n!=0:
        div=n//26;mod=n%26
        re=chr(ord('A')+mod-1 if c==0 else ord('A')+mod)+re
        n=div;c+=1
    return 'A'*(count-len(re))+re

def manage_size(sheet,size,n1,n2):
    if size==None:
        return
    if size[0]!=None:
        for i in range(n2):
            if isinstance(size[0],(tuple,list)):
                if size[0][i]==None:
                    continue
                else:
                    sheet.row_dimensions[i].height=size[0][i]
            else:
                sheet.row_dimensions[i].height=size[0]
    if size[1]!=None:
        for i in range(n1):
            if isinstance(size[1],(tuple,list)):
                if size[1][i]==None:
                    continue
                else:
                    sheet.column_dimensions[ab_order[i]].width=size[1][i]
            else:
                sheet.column_dimensions[ab_order[i]].width=size[1]

'''
data in the .xlsx is recorded in a dictionary.
structure of dictionary------------------------------------------------
key : sheet name
value : the list of data
parameter--------------------------------------------------------------
container : use list/tuple to store the data
by_row : list in value's list is recorded in the row of .xlsx file(if True)
size : size of unit cell
    size={sheetname:(widths,heights)}
        widths, heights could be a number or a list/tuple
'''
def read_xlsx(file,container=list,by_row=True):
    workbook=openpyxl.load_workbook(file)
    sheets=workbook.sheetnames
    ns=len(sheets)
    d={}
    if by_row in (True,False):
        by_row=(by_row,)*ns
    for k in range(ns):
        sheet=workbook[sheets[k]]
        n1=sheet.max_row
        n2=sheet.max_column
        if by_row[k]:
            d[sheet.title]=container(container(sheet.cell(row=i+1,
                                                          column=j+1).value
                                               for j in range(n2))
                                     for i in range(n1))
        else:
            d[sheet.title]=container(container(sheet.cell(row=i+1,
                                                          column=j+1).value
                                               for i in range(n1))
                                     for j in range(n2))
    return d

def write_xlsx(d,output,by_row=True,size=None):
    workbook=openpyxl.Workbook()
    sheets=[]
    names=list(d.keys())
    ns=len(names)
    if by_row in (True,False):
        by_row=(by_row,)*ns
    for k in range(ns):
        sheet=names[k]
        lis=d[sheet]
        # if only create_sheet(), the first empty 'Sheet' is quite annoying
        sheets.append(workbook.active if k==0 else workbook.create_sheet())
        sheets[-1].title=sheet
        if by_row[k]:
            n1=len(d[sheet])
            for i in range(n1):
                n2=len(lis[i])
                for j in range(n2):
                    sheets[-1].cell(i+1,j+1,lis[i][j])
            # manage size
            if size==None:
                continue
            elif isinstance(size,(tuple,list)):
                if len(size)==2:
                    manage_size(sheets[-1],size,n1,max(len(j) for j in lis))
            else:
                if sheet in size.keys():
                    manage_size(sheets[-1],size[sheet],n1,max(len(j) for j in lis))
        else:
            n1=len(d[sheet])
            for i in range(n1):
                n2=len(lis[i])
                for j in range(n2):
                    sheets[-1].cell(j+1,i+1,lis[i][j])
            # manage size
            if size==None:
                continue
            elif isinstance(size,(tuple,list)):
                if len(size)==2:
                    manage_size(sheets[-1],size,max(len(j) for j in lis),n1)
            else:
                if sheet in size.keys():
                    manage_size(sheets[-1],size[sheet],max(len(j) for j in lis),n1)
    workbook.save(filename=output)

def add_xlsx(d,file,by_row=True,size=None):
    if os.path.exists(file):
        workbook=openpyxl.load_workbook(file)
    else:
        workbook=openpyxl.Workbook()
    sheets=[]
    names=list(d.keys())
    ns=len(names)
    if by_row in (True,False):
        by_row=(by_row,)*ns
    for k in range(ns):
        sheet=names[k]
        lis=d[sheet]
        # if only create_sheet(), the first empty 'Sheet' is quite annoying
        sheets.append(workbook.active if k==0 and not os.path.exists(file) else workbook.create_sheet())
        sheets[-1].title=sheet
        if by_row[k]:
            n1=len(d[sheet])
            for i in range(n1):
                n2=len(lis[i])
                for j in range(n2):
                    sheets[-1].cell(i+1,j+1,lis[i][j])
            # manage size
            if size==None:
                continue
            elif isinstance(size,(tuple,list)):
                if len(size)==2:
                    manage_size(sheets[-1],size,n1,max(len(j) for j in lis))
            else:
                if sheet in size.keys():
                    manage_size(sheets[-1],size[sheet],n1,max(len(j) for j in lis))
        else:
            n1=len(d[sheet])
            for i in range(n1):
                n2=len(lis[i])
                for j in range(n2):
                    sheets[-1].cell(j+1,i+1,lis[i][j])
            # manage size
            if size==None:
                continue
            elif isinstance(size,(tuple,list)):
                if len(size)==2:
                    manage_size(sheets[-1],size,max(len(j) for j in lis),n1)
            else:
                if sheet in size.keys():
                    manage_size(sheets[-1],size[sheet],max(len(j) for j in lis),n1)
    workbook.save(filename=file)

def select_xlsx(file,sheet_select,container=list,by_row=True):
    workbook=openpyxl.load_workbook(file)
    sheets=workbook.sheetnames
    ns=len(sheets)
    d={}
    if isinstance(sheet_select,str):
        sheet_select=(sheet_select,)
    if by_row in (True,False):
        by_row=(by_row,)*ns
    for k in range(ns):
        sheet=workbook[sheets[k]]
        if sheet.title not in sheet_select:
            continue
        n1=sheet.max_row
        n2=sheet.max_column
        if by_row[k]:
            d[sheet.title]=container(container(sheet.cell(row=i+1,
                                                          column=j+1).value
                                               for j in range(n2))
                                     for i in range(n1))
        else:
            d[sheet.title]=container(container(sheet.cell(row=i+1,
                                                          column=j+1).value
                                               for i in range(n1))
                                     for j in range(n2))
    return d

__all__=['read_xlsx','write_xlsx','add_xlsx','select_xlsx',]
